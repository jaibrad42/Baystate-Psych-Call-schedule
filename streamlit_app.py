"""
Baystate Psychiatry Call Scheduler — Streamlit Web App
Deploy free at: https://streamlit.io/cloud
"""

import streamlit as st
import json, os, re, calendar, io
from datetime import date, timedelta
from copy import deepcopy

st.set_page_config(
    page_title="Baystate Call Scheduler",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────
# CONFIG — persisted to GitHub repo via API when GITHUB_TOKEN
# secret is set. Falls back to /tmp cache, then local file.
# Setup: Streamlit Cloud → App settings → Secrets → add GITHUB_TOKEN
# ─────────────────────────────────────────────────────────────

import urllib.request, base64

GITHUB_REPO   = "jaibrad42/Baystate-Psych-Call-schedule"
GITHUB_FILE   = "config.json"
GITHUB_BRANCH = "main"
DEFAULT_CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")
TMP_CONFIG_PATH = "/tmp/baystate_config_override.json"

def _gh_token():
    try:
        return st.secrets.get("GITHUB_TOKEN", "")
    except Exception:
        return ""

def _gh_fetch():
    """Fetch config.json content + SHA from GitHub API."""
    token = _gh_token()
    if not token:
        return None, None
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}?ref={GITHUB_BRANCH}"
    req = urllib.request.Request(url, headers={
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "User-Agent": "BaystateScheduler/1.0"
    })
    try:
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode())
            content = base64.b64decode(data["content"].replace("\n","")).decode()
            return json.loads(content), data["sha"]
    except Exception:
        return None, None

def _gh_push(cfg, sha):
    """Commit updated config.json to GitHub."""
    token = _gh_token()
    if not token:
        return False
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"
    body = json.dumps({
        "message": "chore: auto-save config from Baystate Scheduler",
        "content": base64.b64encode(json.dumps(cfg, indent=2).encode()).decode(),
        "sha": sha,
        "branch": GITHUB_BRANCH,
    }).encode()
    req = urllib.request.Request(url, data=body, method="PUT", headers={
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
        "User-Agent": "BaystateScheduler/1.0"
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read().decode())
            st.session_state["_cfg_sha"] = result["content"]["sha"]
            return True
    except Exception:
        return False

def _load_default_config():
    # 1. Try GitHub API (permanent cross-session persistence)
    cfg, sha = _gh_fetch()
    if cfg is not None:
        st.session_state["_cfg_sha"] = sha
        return cfg
    # 2. Try /tmp cache (survives page refresh within container)
    if os.path.exists(TMP_CONFIG_PATH):
        try:
            with open(TMP_CONFIG_PATH) as f:
                return json.load(f)
        except Exception:
            pass
    # 3. Fall back to bundled config.json
    if os.path.exists(DEFAULT_CONFIG_PATH):
        with open(DEFAULT_CONFIG_PATH) as f:
            return json.load(f)
    return {"residents": [], "holidays": [], "program_no_call": [], "no_call_requests": []}

def get_cfg():
    if "config" not in st.session_state:
        st.session_state.config = _load_default_config()
    return st.session_state.config

def save_cfg(cfg):
    st.session_state.config = cfg
    # Write to /tmp for fast in-session persistence
    try:
        with open(TMP_CONFIG_PATH, "w") as f:
            json.dump(cfg, f, indent=2)
    except Exception:
        pass
    # Push to GitHub for permanent cross-session persistence
    sha = st.session_state.get("_cfg_sha")
    if sha:
        _gh_push(cfg, sha)
    else:
        _, sha = _gh_fetch()
        if sha:
            st.session_state["_cfg_sha"] = sha
            _gh_push(cfg, sha)

# ─────────────────────────────────────────────────────────────
# RESIDENT HELPERS
# ─────────────────────────────────────────────────────────────
BLOCKED_ROTATIONS = {"Medicine Wards", "Emergency Medicine", "Neurology"}

def is_off_service(res_id, d, cfg):
    """Return True if resident is off service (away or on a blocking rotation) on weekday d."""
    for r in cfg["residents"]:
        if r["id"] != res_id: continue
        # New unified off_service field
        for o in r.get("off_service", []):
            try:
                if date.fromisoformat(o["start"]) <= d <= date.fromisoformat(o["end"]):
                    return d.weekday() < 5
            except (KeyError, ValueError): pass
        # Backward compat: old away_periods field
        for p in r.get("away_periods", []):
            try:
                if date.fromisoformat(p["start"]) <= d <= date.fromisoformat(p["end"]):
                    return d.weekday() < 5
            except (KeyError, ValueError): pass
        # Backward compat: old blocked_rotations field
        for b in r.get("blocked_rotations", []):
            try:
                if date.fromisoformat(b["start"]) <= d <= date.fromisoformat(b["end"]):
                    return d.weekday() < 5
            except (KeyError, ValueError): pass
    return False

def is_blocking_rotation(res_id, d, cfg):
    """Return True if resident is on a call-blocking rotation (Medicine Wards, EM, Neurology)."""
    for r in cfg["residents"]:
        if r["id"] != res_id: continue
        for o in r.get("off_service", []):
            try:
                if (date.fromisoformat(o["start"]) <= d <= date.fromisoformat(o["end"])
                        and o.get("reason","") in BLOCKED_ROTATIONS):
                    return True
            except (KeyError, ValueError): pass
        for b in r.get("blocked_rotations", []):
            try:
                if (date.fromisoformat(b["start"]) <= d <= date.fromisoformat(b["end"])
                        and b.get("name","") in BLOCKED_ROTATIONS):
                    return True
            except (KeyError, ValueError): pass
    return False

def get_holiday(d, cfg):
    dk = d.isoformat()
    for h in cfg["holidays"]:
        for entry in h["entries"]:
            if entry["date"] == dk:
                return {"name": h["name"], "aptu": entry.get("aptu",""), "consult": entry.get("consult","")}
    return None

def holiday_soon(res_id, d, cfg, within=3):
    for i in range(1, within+1):
        h = get_holiday(d + timedelta(days=i), cfg)
        if h and (h["aptu"] == res_id or h["consult"] == res_id):
            return True
    return False

def is_no_call(res_id, d, cfg):
    dk = d.isoformat()
    return any(e["resident"] == res_id and e["date"] == dk
               for e in cfg.get("no_call_requests", []))

def is_program_nc(d, cfg):
    return d.isoformat() in cfg.get("program_no_call", [])

def intern_limits(month, year):
    am = (year - 2026) * 12 + (month - 7) + 1
    am = max(am, 1)
    wd = 2
    we = 0 if am == 1 else (1 if am <= 3 else 2)
    solo_we = am >= 2
    return wd, we, solo_we

SOFT_CAPS = {1: 2, 2: 5, 3: 3, 4: 2}
HARD_CAPS = {1: 3, 2: 6, 3: 5, 4: 4}

class State:
    def __init__(self, cfg):
        rs = active_residents(cfg)
        self.last  = {r["id"]: None for r in rs}
        self.month = {r["id"]: 0    for r in rs}
        self.wf    = {r["id"]: 0    for r in rs}
        self.iwd   = {r["id"]: 0    for r in rs if r["pgy"] == 1}
        self.we_sats = {r["id"]: set() for r in rs}
        self._current_month = None

    def reset_month_counters(self, year, month):
        if self._current_month != (year, month):
            for k in self.iwd: self.iwd[k] = 0
            self._current_month = (year, month)

    def eligible(self, res_id, d, role, cfg):
        if is_off_service(res_id, d, cfg): return False
        if is_no_call(res_id, d, cfg): return False
        if is_program_nc(d, cfg): return False
        if holiday_soon(res_id, d, cfg): return False
        rb = res_by_id(cfg)
        if res_id not in rb: return False
        p = rb[res_id]["pgy"]
        if role == "aptu_wd":      return p >= 2
        if role == "aptu_we_jul":  return p >= 2 and not self._over_consec(res_id, d)
        if role == "aptu_we_aug":
            if p == 1 and is_blocking_rotation(res_id, d, cfg): return False
            return not self._over_consec(res_id, d)
        if role == "consult":
            return p >= 3 and not self._over_consec(res_id, d)
        if role == "intern":
            return p == 1 and not is_blocking_rotation(res_id, d, cfg)
        return False

    def _over_consec(self, res_id, d):
        sat = d if d.weekday() == 5 else d - timedelta(days=(d.weekday()-5)%7)
        w = self.we_sats[res_id]
        return (sat-timedelta(7)) in w and (sat-timedelta(14)) in w and (sat-timedelta(21)) in w

    def score(self, res_id, d, role, cfg):
        last = self.last[res_id]
        gap  = (d - last).days if last else 99
        if gap < 2: return 1e9
        q = 0 if gap >= 4 else (100 if gap == 3 else 500)
        rb = res_by_id(cfg)
        pgy = rb[res_id]["pgy"] if res_id in rb else 2
        cap = SOFT_CAPS.get(pgy, 5)
        cnt_pen = self.month[res_id] * (100 / max(cap, 1))
        wf_pen  = self.wf[res_id] * 12 if role == "aptu_wd" and d.weekday() in (2,4) else 0
        return q + cnt_pen + wf_pen - pgy * 0.5

    def best(self, pool, d, role, cfg, exclude=()):
        cands = [r for r in pool if r not in exclude and self.eligible(r, d, role, cfg)]
        if not cands: return None
        cands.sort(key=lambda r: self.score(r, d, role, cfg))
        return cands[0] if self.score(cands[0], d, role, cfg) < 1e8 else None

    def fallback(self, pool, d, role, cfg, exclude=()):
        cands = [r for r in pool if r not in exclude and self.eligible(r, d, role, cfg)
                 and (self.last[r] is None or (d - self.last[r]).days >= 2)]
        if cands:
            cands.sort(key=lambda r: -(d-self.last[r]).days if self.last[r] else -99)
            return cands[0]
        cands2 = [r for r in pool if r not in exclude and self.eligible(r, d, role, cfg)
                  and (self.last[r] is None or (d - self.last[r]).days >= 1)]
        if cands2:
            cands2.sort(key=lambda r: self.month[r])
            return cands2[0]
        return None

    def record(self, res_id, d, role):
        if res_id not in self.last: return
        self.last[res_id] = d
        self.month[res_id] += 1
        if role == "aptu_wd" and d.weekday() in (2,4):
            self.wf[res_id] += 1
        if d.weekday() >= 5 and role in ("aptu_we_jul","aptu_we_aug","consult"):
            sat = d if d.weekday()==5 else d - timedelta(days=(d.weekday()-5)%7)
            self.we_sats[res_id].add(sat)
        if res_id in self.iwd and role == "intern_wd":
            self.iwd[res_id] += 1


def schedule_month(year, month, state, cfg):
    pools = pgy_pools(cfg)
    first = date(year, month, 1)
    last  = date(year, month+1, 1) - timedelta(1) if month < 12 else date(year+1,1,1) - timedelta(1)
    days  = [first + timedelta(i) for i in range((last-first).days+1)]
    iwd_lim, iwe_lim, solo_we = intern_limits(month, year)
    state.reset_month_counters(year, month)

    for d in days:
        h = get_holiday(d, cfg)
        if h:
            if h["aptu"]:    state.month[h["aptu"]]    = state.month.get(h["aptu"],0)    + 1
            if h["consult"]: state.month[h["consult"]] = state.month.get(h["consult"],0) + 1

    sched = {}; warnings = []; prev_intern = None

    for d in days:
        dk = d.isoformat(); dow = d.weekday(); is_wk = dow <= 4

        if is_program_nc(d, cfg):
            sched[dk] = {"type": "no_call"}; continue

        h = get_holiday(d, cfg)
        if h:
            entry_h = {"type":"holiday","name":h["name"]}
            if h["aptu"]:    entry_h["aptu"]    = h["aptu"]
            if h["consult"]: entry_h["consult"] = h["consult"]
            sched[dk] = entry_h
            if h["aptu"]:    state.record(h["aptu"],    d, "aptu_wd" if is_wk else "aptu_we_jul")
            if h["consult"]: state.record(h["consult"], d, "consult")
            prev_intern = None; continue

        entry = {"type": "weekday" if is_wk else "weekend"}

        if not is_wk:
            aptu_role = "aptu_we_aug" if solo_we else "aptu_we_jul"
            aptu_pool = pools["all"] if solo_we else pools["upper"]
            con = state.best(pools["pgy34"], d, "consult", cfg) or \
                  state.fallback(pools["pgy34"], d, "consult", cfg)
            if con is None:
                warnings.append(f"{dk}: NO eligible Consult — UNCOVERED")
            elif state.last.get(con) and (d - state.last[con]).days < 2:
                warnings.append(f"{dk}: Consult {con} back-to-back — needs review")
            aptu = state.best([r for r in aptu_pool if r != con], d, aptu_role, cfg) or \
                   state.fallback([r for r in aptu_pool if r != con], d, aptu_role, cfg)
            if aptu is None:
                warnings.append(f"{dk}: NO eligible APTU — UNCOVERED")
            elif state.last.get(aptu) and (d - state.last[aptu]).days < 2:
                warnings.append(f"{dk}: APTU {aptu} back-to-back — needs review")
            if con:  entry["consult"] = con;  state.record(con,  d, "consult")
            if aptu: entry["aptu"]    = aptu; state.record(aptu, d, aptu_role)
            prev_intern = None

        else:
            aptu = state.best(pools["upper"], d, "aptu_wd", cfg) or \
                   state.fallback(pools["upper"], d, "aptu_wd", cfg)
            if aptu is None:
                warnings.append(f"{dk}: NO eligible APTU — UNCOVERED")
            else:
                gap = (d - state.last[aptu]).days if state.last.get(aptu) else 99
                if gap == 2:
                    warnings.append(f"{dk}: APTU {aptu} at q3 — needs review")
            intern = None
            rb = res_by_id(cfg)
            if aptu and rb.get(aptu,{}).get("pgy",0) >= 3 and iwd_lim > 0:
                for r in pools["interns"]:
                    if state.iwd.get(r,0) < iwd_lim and state.eligible(r, d, "intern", cfg) and r != prev_intern:
                        intern = r; break
            if aptu:   entry["aptu"]   = aptu;   state.record(aptu, d, "aptu_wd")
            if intern: entry["intern"] = intern; state.record(intern, d, "intern_wd")
            prev_intern = intern

        sched[dk] = entry
    return sched, warnings


def schedule_jeopardy(sched, cfg):
    rb = res_by_id(cfg)
    for dk, entry in sched.items():
        if entry.get("type") == "no_call": continue
        d = date.fromisoformat(dk)
        assigned = {entry.get(k) for k in ("aptu","consult","intern") if entry.get(k)}
        cands = [r["id"] for r in active_residents(cfg)
                 if r["pgy"] >= 3 and r["id"] not in assigned
                 and not is_off_service(r["id"], d, cfg) and not is_no_call(r["id"], d, cfg)]
        if cands:
            def month_load(rid):
                return sum(1 for dk2,e2 in sched.items()
                           if date.fromisoformat(dk2).month == d.month
                           and e2.get("jeopardy") == rid)
            cands.sort(key=lambda r: (month_load(r), r))
            entry["jeopardy"] = cands[0]
        else:
            entry["jeopardy"] = "UNCOVERED"
    return sched


# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────
def add_reference_sheet(wb, cfg):
    """Add a Reference sheet listing all locked holidays, program no-call days, and individual no-call requests."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl import Workbook
    import copy
    from collections import defaultdict

    ws = wb.create_sheet("Reference")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    C_HDR = "1E3A8A"; C_HOL = "FEE2E2"; C_PRG = "DBEAFE"
    DAYS_ABB = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]

    def rfill(h): return PatternFill("solid", fgColor=h)
    def rfont(bold=False, sz=10, color="111827"):
        return Font(name="Arial", bold=bold, size=sz, color=color)
    def rborder():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)
    def ralign(h="left"):
        return Alignment(horizontal=h, vertical="center")

    residents = cfg.get("residents", [])
    active_res = [r for r in residents if r.get("active", True)]
    SHORT = {r["id"]: r["full"] for r in residents}
    pgy_map = {r["id"]: r["pgy"] for r in residents}
    pgy_fill_map = {1:"FEF3C7", 2:"DBEAFE", 3:"EDE9FE", 4:"D1FAE5"}

    row = 1

    # Section 1: Holiday Assignments
    for c, h in enumerate(["Holiday","Date","Day","APTU","Consult"], 1):
        cell = ws.cell(row, c, h)
        cell.fill = rfill(C_HDR); cell.font = rfont(bold=True, sz=11, color="FFFFFF")
        cell.alignment = ralign("center"); cell.border = rborder()
    row += 1

    for hol in cfg.get("holidays", []):
        for i, entry in enumerate(hol["entries"]):
            try:
                d = date.fromisoformat(entry["date"])
            except:
                continue
            aptu_id = entry.get("aptu","")
            con_id  = entry.get("consult","")
            aptu_name = SHORT.get(aptu_id, aptu_id) if aptu_id else "—"
            con_name  = SHORT.get(con_id,  con_id)  if con_id  else "—"
            vals = [hol["name"] if i==0 else "", d.strftime("%b %d, %Y"),
                    DAYS_ABB[d.weekday()], aptu_name, con_name]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row, c, v)
                cell.fill = rfill(C_HOL); cell.font = rfont(sz=10)
                cell.alignment = ralign("center" if c > 1 else "left")
                cell.border = rborder()
            row += 1
        row += 1

    # Section 2: Program No-Call Days
    row += 1
    ws.cell(row, 1, "Program No-Call Days").font = rfont(bold=True, sz=12)
    row += 1
    for c, h in enumerate(["Date","Day","Note"], 1):
        cell = ws.cell(row, c, h)
        cell.fill = rfill(C_HDR); cell.font = rfont(bold=True, sz=11, color="FFFFFF")
        cell.alignment = ralign("center"); cell.border = rborder()
    row += 1
    for ds in sorted(cfg.get("program_no_call",[])):
        try:
            d = date.fromisoformat(ds)
        except:
            continue
        for c, v in enumerate([d.strftime("%b %d, %Y"), DAYS_ABB[d.weekday()], "No call scheduled"], 1):
            cell = ws.cell(row, c, v)
            cell.fill = rfill(C_PRG); cell.font = rfont(sz=10)
            cell.alignment = ralign("center" if c > 1 else "left"); cell.border = rborder()
        row += 1

    # Section 3: Individual No-Call Requests
    row += 2
    ws.cell(row, 1, "Individual No-Call Requests").font = rfont(bold=True, sz=12)
    row += 1
    ws.column_dimensions["D"].width = 60
    for c, h in enumerate(["Resident","PGY","Month","Dates"], 1):
        cell = ws.cell(row, c, h)
        cell.fill = rfill(C_HDR); cell.font = rfont(bold=True, sz=11, color="FFFFFF")
        cell.alignment = ralign("center"); cell.border = rborder()
    row += 1

    nc_by_res = defaultdict(list)
    for entry in cfg.get("no_call_requests", []):
        nc_by_res[entry["resident"]].append(entry["date"])

    for r in active_res:
        dates = sorted(nc_by_res.get(r["id"], []))
        if not dates:
            continue
        by_month = defaultdict(list)
        for ds in dates:
            by_month[ds[:7]].append(ds)
        first_row = row
        pf = rfill(pgy_fill_map.get(r["pgy"],"FFFFFF"))
        for ym in sorted(by_month.keys()):
            date_strs = []
            for ds in by_month[ym]:
                try:
                    date_strs.append(date.fromisoformat(ds).strftime("%b %-d"))
                except:
                    date_strs.append(ds)
            ws.cell(row, 1, r["full"] if row == first_row else "").font = rfont(sz=10)
            ws.cell(row, 1).fill = copy.copy(pf); ws.cell(row, 1).border = rborder()
            ws.cell(row, 2, f"PGY-{r['pgy']}" if row == first_row else "").font = rfont(sz=10)
            ws.cell(row, 2).fill = copy.copy(pf); ws.cell(row, 2).border = rborder()
            ws.cell(row, 2).alignment = ralign("center")
            try:
                label = date.fromisoformat(by_month[ym][0]).strftime("%b %Y")
            except:
                label = ym
            ws.cell(row, 3, label).font = rfont(sz=10)
            ws.cell(row, 3).fill = copy.copy(pf); ws.cell(row, 3).border = rborder()
            ws.cell(row, 3).alignment = ralign("center")
            ws.cell(row, 4, ",  ".join(date_strs)).font = rfont(sz=10)
            ws.cell(row, 4).fill = copy.copy(pf); ws.cell(row, 4).border = rborder()
            ws.cell(row, 4).alignment = Alignment(horizontal="left", wrap_text=True)
            ws.row_dimensions[row].height = 15
            row += 1
        if row > first_row:
            row += 1


def export_xlsx_bytes(all_scheds, all_warnings, cfg):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import copy

    wb = Workbook()
    wb.remove(wb.active)
    SHORT = {r["id"]: r["full"] for r in cfg["residents"]}

    C_HDR="1E3A8A"; C_WKND="EDE9FE"; C_HOL="FEE2E2"; C_GRN="D1FAE5"
    C_WF="DBEAFE"; C_UNC="FCA5A5"; C_AMBI="FEF9C3"; C_WHT="FFFFFF"
    C_SAT="F3F0FF"; C_SUN="EDE9FE"; C_DAYN="F8FAFC"; DAYS_WIDE=24

    def fill(h): return PatternFill("solid", fgColor=h)
    def font(bold=False, sz=10, color="111827", italic=False):
        return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)
    def border(style="thin", color="D1D5DB"):
        s=Side(style=style,color=color); return Border(left=s,right=s,top=s,bottom=s)
    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    ROWS_PER_WEEK=5; ROW_H_DAY=15; ROW_H_ROLE=18

    combined = {r["id"]:{"total":0,"wk":0,"we":0,"wf":0} for r in cfg["residents"]}
    monthly_counts = {r["id"]: [] for r in cfg["residents"]}

    for (year, month), sched in all_scheds.items():
        warnings = all_warnings.get((year,month),[])
        mlabel   = date(year,month,1).strftime("%b %Y")
        ws = wb.create_sheet(mlabel)
        ws.merge_cells("A1:G1")
        ws["A1"].value = date(year,month,1).strftime("%B %Y").upper() + \
                         "  —  BAYSTATE PSYCHIATRY CALL SCHEDULE"
        ws["A1"].font = font(bold=True,sz=13,color="FFFFFF")
        ws["A1"].fill = fill(C_HDR); ws["A1"].alignment = align()
        ws.row_dimensions[1].height = 26

        for col, day in enumerate(["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"],1):
            c = ws.cell(2,col,day)
            c.font=font(bold=True,sz=11,color="FFFFFF"); c.fill=fill(C_HDR)
            c.alignment=align(); c.border=border()
            ws.column_dimensions[get_column_letter(col)].width = DAYS_WIDE
        ws.row_dimensions[2].height = 20

        first_day  = date(year,month,1); start_col = first_day.weekday()
        num_days   = calendar.monthrange(year,month)[1]; current_row = 3
        total_weeks = ((num_days+start_col-1)//7)+1

        for week_i in range(total_weeks):
            r0 = current_row + week_i * ROWS_PER_WEEK
            ws.row_dimensions[r0].height = ROW_H_DAY
            for rr in range(1,ROWS_PER_WEEK): ws.row_dimensions[r0+rr].height = ROW_H_ROLE

        for day_num in range(1, num_days+1):
            d=date(year,month,day_num); dow=d.weekday(); col=dow+1
            slot=day_num-1+start_col; week_i=slot//7; r0=current_row+week_i*ROWS_PER_WEEK
            e=sched.get(d.isoformat(),{}); t=e.get("type",""); wf=dow in (2,4) and t=="weekday"

            if   t=="holiday":   bg=C_HOL
            elif t=="no_call":   bg=C_AMBI
            elif dow==5:         bg=C_SAT
            elif dow==6:         bg=C_SUN
            elif wf:             bg=C_WF
            else:                bg=C_WHT

            dn=ws.cell(r0,col); dn.value=day_num
            dn.font=font(bold=True,sz=11,color="374151")
            dn.fill=fill(C_DAYN if bg==C_WHT else bg)
            dn.alignment=align(h="right"); dn.border=border()
            if t=="holiday":
                dn.value=f"{day_num}  🔒 {e.get('name','')}"; dn.font=font(bold=True,sz=9,color="991B1B")
            elif t=="no_call":
                dn.value=f"{day_num}  — No call"; dn.font=font(bold=True,sz=9,color="92400E",italic=True)

            roles_map={"weekday":[("aptu","APTU"),("consult",""),("intern","Intern")],
                       "weekend":[("aptu","APTU"),("consult","Consult"),("","")],
                       "holiday":[("aptu","APTU"),("consult","Consult"),("","")],
                       "no_call":[("",""),("",""),("","")]}
            roles=roles_map.get(t,[("",""),("",""),("","")])
            for ri,(key,label) in enumerate(roles):
                rc=r0+1+ri; rid=e.get(key,"") if key else ""; name=SHORT.get(rid,rid) if rid else ""
                cell=ws.cell(rc,col); cell.border=border(); cell.alignment=align(h="left",wrap=True)
                if not label or (t=="weekday" and ri==1):
                    cell.fill=fill(bg); cell.value=""
                elif name=="" and ri<2 and t not in ("no_call",):
                    cell.value="UNCOVERED"; cell.font=font(bold=True,sz=9,color="7F1D1D"); cell.fill=fill(C_UNC)
                else:
                    cell.value=f"  {name}" if name else ""
                    if ri==2 and name:
                        cell.fill=fill(C_GRN); cell.font=font(sz=9,color="065F46")
                    else:
                        cell.fill=fill(bg); cell.font=font(sz=9,color="1E3A8A" if name else "9CA3AF",bold=bool(name))

            jrc=r0+4; jname=SHORT.get(e.get("jeopardy",""),e.get("jeopardy",""))
            jcell=ws.cell(jrc,col); jcell.border=border(); jcell.alignment=align(h="left",wrap=True)
            if t=="no_call" or not e.get("jeopardy"):
                jcell.fill=fill(bg); jcell.value=""
            elif jname=="UNCOVERED":
                jcell.value="  Jep: UNCOVERED"; jcell.fill=fill(C_UNC); jcell.font=font(bold=True,sz=9,color="7F1D1D")
            else:
                jcell.value=f"  Jep: {jname}"; jcell.fill=PatternFill("solid",fgColor="FFF7CD")
                jcell.font=font(sz=9,color="854D0E",bold=True)

        for week_i in range(total_weeks):
            r0=current_row+week_i*ROWS_PER_WEEK
            for col in range(1,8):
                for rr in range(ROWS_PER_WEEK):
                    cell=ws.cell(r0+rr,col)
                    if cell.value is None:
                        cell.fill=fill("F8FAFC"); cell.border=border(color="E2E8F0")

        # Warnings sheet
        ws_w=wb.create_sheet(f"{mlabel[:3]} Warnings")
        ws_w.column_dimensions["A"].width=80
        ws_w.cell(1,1,f"Warnings — {mlabel}").font=font(bold=True,sz=12)
        if not warnings: ws_w.cell(2,1,"✓  No constraint violations.").font=font(sz=11,color="065F46")
        else:
            for i,w in enumerate(warnings,2):
                ws_w.cell(i,1,f"•  {w}").font=font(sz=10,color="7F1D1D" if "UNCOVERED" in w or "back-to-back" in w else "92400E")

        # Accumulate
        _mc = {r["id"]: 0 for r in cfg["residents"]}
        for dk,ev in sched.items():
            d=date.fromisoformat(dk)
            if d.month!=month: continue
            is_we=d.weekday()>=5 or ev.get("type")=="holiday"
            for role in ["aptu","consult","intern"]:
                rid=ev.get(role)
                if rid and rid in combined:
                    combined[rid]["total"]+=1
                    _mc[rid]+=1
                    if is_we: combined[rid]["we"]+=1
                    else: combined[rid]["wk"]+=1
            aptu=ev.get("aptu")
            if aptu and not is_we and d.weekday() in (2,4): combined[aptu]["wf"]+=1
        for rid in monthly_counts: monthly_counts[rid].append(_mc.get(rid,0))
    # All Counts sheet
    ws_c=wb.create_sheet("All Counts")
    hdrs=["Resident","PGY","Total Calls","Weekday","Weekend","Wed/Fri APTU","Status"]
    for c,h in enumerate(hdrs,1):
        cell=ws_c.cell(1,c,h); cell.fill=fill(C_HDR)
        cell.font=font(bold=True,sz=11,color="FFFFFF"); cell.alignment=align(); cell.border=border()
    ws_c.column_dimensions["A"].width=24
    for col in "BCDEF": ws_c.column_dimensions[col].width=14
    ws_c.column_dimensions["G"].width=22
    pgy_fill={1:"FEF3C7",2:"DBEAFE",3:"EDE9FE",4:"D1FAE5"}
    for i,r in enumerate(cfg["residents"],2):
        rf=fill(pgy_fill.get(r["pgy"],"FFFFFF"))
        for c,val in enumerate([r["full"],f"PGY-{r['pgy']}",
                                 combined[r["id"]]["total"],combined[r["id"]]["wk"],
                                 combined[r["id"]]["we"],combined[r["id"]]["wf"]],1):
            cell=ws_c.cell(i,c,val); cell.fill=copy.copy(rf); cell.font=font(sz=10)
            cell.alignment=align(h="left" if c==1 else "center"); cell.border=border()
        # Status column (max calls in any single month vs monthly cap)
        sc2 = SOFT_CAPS.get(r["pgy"], 5)
        hc2 = HARD_CAPS.get(r["pgy"], 6)
        mo_counts = monthly_counts.get(r["id"], [0])
        max_mo = max(mo_counts) if mo_counts else 0
        if max_mo > hc2:   status_txt=f"OVER HARD CAP ({max_mo}/{hc2}/mo)"; sf="FCA5A5"; sbold=True
        elif max_mo > sc2: status_txt=f"Over soft cap ({max_mo}/{sc2}/mo)"; sf="FDE68A"; sbold=False
        else:               status_txt=f"OK ✓ ({combined[r['id']]['total']} calls)"; sf="A7F3D0"; sbold=False
        scell = ws_c.cell(i, 7, status_txt)
        scell.fill = fill(sf); scell.font = font(bold=sbold, sz=10)
        scell.alignment = align(); scell.border = border()

    # Move All Counts to front
    add_reference_sheet(wb, cfg)
    wb.move_sheet("All Counts", offset=-len(wb.sheetnames)+1)
    wb.move_sheet("Reference", offset=-len(wb.sheetnames)+2)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stSidebar"] { background: #1C1C1E; }
[data-testid="stSidebar"] * { color: #F2F2F7 !important; }
.stTabs [data-baseweb="tab-list"] { gap: 4px; }
.stTabs [data-baseweb="tab"] { background: #2C2C2E; border-radius: 6px; padding: 6px 16px; color: #8E8E93; }
.stTabs [aria-selected="true"] { background: #3B82F6 !important; color: white !important; }
.cal-grid { display: grid; grid-template-columns: repeat(7, 1fr); gap: 3px; margin-top: 8px; }
.cal-header-row { display: grid; grid-template-columns: repeat(7, 1fr); gap: 3px; margin-top: 12px; }
.dow-hdr { text-align: center; font-size: 11px; font-weight: 600; color: #8E8E93; padding: 4px; }
.cal-cell { background: #2C2C2E; border-radius: 6px; padding: 6px 8px; min-height: 78px; font-size: 11px; }
.cal-cell.weekend { background: #26263A; }
.cal-cell.holiday { background: #3B1515; border: 1px solid #7F1D1D; }
.cal-cell.no-call  { background: #2A2A1A; }
.cal-cell.wf       { border-left: 2px solid #3B82F6; }
.cal-cell.empty    { background: transparent; min-height: 0; }
.day-num { font-weight: 700; color: #8E8E93; margin-bottom: 3px; }
.pill { display: inline-block; padding: 1px 6px; border-radius: 3px; font-size: 10px; font-weight: 500; margin: 1px 0; width: 100%; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.pill-ul      { background: #1e3a8a33; color: #93C5FD; }
.pill-aptu    { background: #3d247333; color: #C4B5FD; }
.pill-consult { background: #06574633; color: #6EE7B7; }
.pill-intern  { background: #7c2d1233; color: #FB923C; font-weight: 600; }
.pill-hol     { background: #7f1d1d55; color: #FCA5A5; }
.pill-jep     { background: #06474733; color: #2DD4BF; font-style: italic; }
.pill-uncov   { background: #7f1d1d; color: white; font-weight: 700; }
.role-lbl     { font-size: 9px; color: #555; margin-bottom: 1px; }
.warn-hard    { background: #3B1515; border-left: 3px solid #EF4444; padding: 6px 10px; border-radius: 4px; font-family: monospace; font-size: 12px; margin-bottom: 4px; color: #FCA5A5; }
.warn-soft    { background: #2A1F0A; border-left: 3px solid #F59E0B; padding: 6px 10px; border-radius: 4px; font-family: monospace; font-size: 12px; margin-bottom: 4px; color: #FCD34D; }
.warn-ok      { background: #0A2A1A; border-left: 3px solid #10B981; padding: 6px 10px; border-radius: 4px; font-size: 12px; color: #6EE7B7; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📅 Baystate Scheduler")
    st.markdown("---")

    # Config upload/download
    with st.expander("💾 Config file", expanded=True):
        st.caption("Download after making changes to save your data.")
        cfg_bytes = json.dumps(get_cfg(), indent=2).encode()
        st.download_button("⬇ Download config.json", cfg_bytes,
                           file_name="config.json", mime="application/json",
                           use_container_width=True)
        up = st.file_uploader("⬆ Upload config.json", type="json", label_visibility="collapsed")
        if up:
            try:
                new_cfg = json.load(up)
                save_cfg(new_cfg)
                st.success("Config loaded!")
                st.rerun()
            except Exception as e:
                st.error(f"Invalid JSON: {e}")

    st.markdown("---")
    st.markdown("**Generate schedule**")
    MONTH_MAP = {"July":7,"August":8,"September":9,"October":10,
                 "November":11,"December":12,"January":1,"February":2,
                 "March":3,"April":4,"May":5,"June":6}
    sel_month  = st.selectbox("Starting month", list(MONTH_MAP.keys()), index=0)
    sel_year   = st.selectbox("Year", [2026, 2027], index=0)
    sel_n      = st.selectbox("# of months", [1,2,3,4,5,6,12], index=5)

    if st.button("▶  Generate Schedule", type="primary", use_container_width=True):
        cfg = get_cfg()
        if not active_residents(cfg):
            st.error("No active residents in config.")
        else:
            with st.spinner("Generating…"):
                state = State(cfg)
                all_scheds = {}; all_warns = {}
                y, m = sel_year, MONTH_MAP[sel_month]
                for _ in range(sel_n):
                    sched, warns = schedule_month(y, m, state, cfg)
                    schedule_jeopardy(sched, cfg)
                    all_scheds[(y,m)] = sched; all_warns[(y,m)] = warns
                    m += 1
                    if m > 12: m = 1; y += 1
                st.session_state.all_scheds = all_scheds
                st.session_state.all_warns  = all_warns
                # Save schedule to config for cross-session persistence
                try:
                    _cfg2 = get_cfg()
                    _cfg2["last_schedule"] = {
                        "scheds": {f"{y}-{m}": v for (y,m),v in all_scheds.items()},
                        "warns":  {f"{y}-{m}": v for (y,m),v in all_warns.items()},
                    }
                    save_cfg(_cfg2)
                except Exception:
                    pass
                total_warns = sum(len(v) for v in all_warns.values())
                if total_warns:
                    st.warning(f"{total_warns} warning(s) — check Warnings tab")
                else:
                    st.success("Done! No warnings.")

    if "all_scheds" in st.session_state and st.session_state.all_scheds:
        cfg = get_cfg()
        xlsx = export_xlsx_bytes(st.session_state.all_scheds, st.session_state.all_warns, cfg)
        months_list = sorted(st.session_state.all_scheds.keys())
        label = date(months_list[0][0],months_list[0][1],1).strftime("%b%Y")
        if len(months_list) > 1:
            label += f"_to_{date(months_list[-1][0],months_list[-1][1],1).strftime('%b%Y')}"
        st.download_button("⬇ Download Excel", xlsx,
                           file_name=f"Baystate_Schedule_{label}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

    if "all_scheds" in st.session_state:
        if st.button("🗑 Clear schedule", use_container_width=True):
            del st.session_state["all_scheds"]
            del st.session_state["all_warns"]
            st.rerun()


# ─────────────────────────────────────────────────────────────
# STARTUP: restore last saved schedule from config
# ─────────────────────────────────────────────────────────────
if "all_scheds" not in st.session_state:
    _saved = get_cfg().get("last_schedule")
    if _saved:
        try:
            # Keys stored as "YYYY-M" strings; restore to (year,month) tuples
            st.session_state.all_scheds = {
                tuple(int(x) for x in k.split("-")): v
                for k, v in _saved.get("scheds", {}).items()
            }
            st.session_state.all_warns = {
                tuple(int(x) for x in k.split("-")): v
                for k, v in _saved.get("warns", {}).items()
            }
        except Exception:
            pass

# ─────────────────────────────────────────────────────────────
# MAIN TABS
# ─────────────────────────────────────────────────────────────
tab_cal, tab_res, tab_hol, tab_nc, tab_pnc = st.tabs([
    "📅 Schedule", "👥 Residents", "🔒 Holidays", "🚫 No-Call Requests", "📵 Program No-Call"
])

# ─── helpers ────────────────────────────────────────────────
def res_display_name(res_id, cfg):
    rb = res_by_id(cfg)
    return rb[res_id]["full"] if res_id in rb else res_id

def render_calendar(sched, cfg, year, month):
    rb = res_by_id(cfg)
    def rname(rid): return rb[rid]["full"].split()[-1] if rid in rb else rid

    first_dow = date(year,month,1).weekday()
    num_days  = calendar.monthrange(year,month)[1]

    html = '<div class="cal-header-row">'
    for d in ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]:
        html += f'<div class="dow-hdr">{d}</div>'
    html += '</div><div class="cal-grid">'

    for slot in range(42):
        day_num = slot - first_dow + 1
        if day_num < 1 or day_num > num_days:
            html += '<div class="cal-cell empty"></div>'
            continue
        d = date(year,month,day_num); dk = d.isoformat()
        e = sched.get(dk, {}); t = e.get("type","")
        dow = d.weekday()
        is_wknd = dow >= 5; is_hol = t=="holiday"; is_nc = t=="no_call"
        is_wf   = dow in (2,4) and t=="weekday"

        cls = "cal-cell"
        if is_hol:   cls += " holiday"
        elif is_nc:  cls += " no-call"
        elif is_wknd:cls += " weekend"
        if is_wf:    cls += " wf"

        dow_abbr = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][dow]
        html += f'<div class="{cls}"><div class="day-num">{day_num} <span style="font-weight:400;font-size:9px;color:#555">{dow_abbr}</span>'
        if is_hol: html += f' <span style="font-size:9px;color:#FCA5A5">🔒</span>'
        if is_wf:  html += f' <span style="font-size:9px;color:#93C5FD">★</span>'
        html += '</div>'

        if is_nc:
            html += '<div style="font-size:10px;color:#8E8E93;font-style:italic">no call</div>'
        elif is_hol:
            aptu = rname(e.get("aptu","")) if e.get("aptu") else "—"
            con  = rname(e.get("consult","")) if e.get("consult") else "—"
            html += f'<div class="pill pill-hol">A: {aptu}</div>'
            html += f'<div class="pill pill-hol">C: {con}</div>'
        elif t == "weekend":
            aptu = e.get("aptu"); con = e.get("consult")
            html += f'<div class="pill {"pill-aptu" if aptu else "pill-uncov"}">'
            html += f'{"A: "+rname(aptu) if aptu else "UNCOV"}</div>'
            html += f'<div class="pill {"pill-consult" if con else "pill-uncov"}">'
            html += f'{"C: "+rname(con) if con else "UNCOV"}</div>'
            if e.get("intern"):
                html += f'<div class="pill pill-intern">I: {rname(e["intern"])}</div>'
            if e.get("jeopardy"):
                html += f'<div class="pill pill-jep">J: {rname(e["jeopardy"])}</div>'
        elif t == "weekday":
            aptu = e.get("aptu")
            html += f'<div class="pill {"pill-ul" if aptu else "pill-uncov"}">'
            html += f'{"UL: "+rname(aptu) if aptu else "UNCOV"}</div>'
            if e.get("intern"):
                html += f'<div class="pill pill-intern">I: {rname(e["intern"])}</div>'
            if e.get("jeopardy"):
                html += f'<div class="pill pill-jep">J: {rname(e["jeopardy"])}</div>'
        html += '</div>'
    html += '</div>'
    return html


# ─── Tab 1: Schedule ─────────────────────────────────────────
with tab_cal:
    if "all_scheds" not in st.session_state or not st.session_state.all_scheds:
        st.info("No schedule generated yet. Use the sidebar to generate one.")
    else:
        cfg = get_cfg()
        months = sorted(st.session_state.all_scheds.keys())
        month_labels = [date(y,m,1).strftime("%B %Y") for y,m in months]
        sel_idx = st.selectbox("Month", range(len(months)), format_func=lambda i: month_labels[i])
        year, month = months[sel_idx]
        sched = st.session_state.all_scheds[(year,month)]

        st.markdown(f"### {month_labels[sel_idx]}")

        # Legend
        cols = st.columns(6)
        legend = [("pill-ul","Weekday UL"),("pill-aptu","Wknd APTU"),
                  ("pill-consult","Consult"),("pill-intern","Intern"),
                  ("pill-hol","Holiday"),("pill-jep","Jeopardy")]
        for col,(cls,lbl) in zip(cols,legend):
            col.markdown(f'<span class="pill {cls}">{lbl}</span>', unsafe_allow_html=True)

        st.markdown(render_calendar(sched, cfg, year, month), unsafe_allow_html=True)

        # Warnings for this month
        warns = st.session_state.all_warns.get((year,month),[])
        if warns:
            st.markdown("---")
            st.markdown("**⚠ Warnings for this month**")
            for w in warns:
                cls = "warn-hard" if ("UNCOVERED" in w or "back-to-back" in w) else "warn-soft"
                st.markdown(f'<div class="{cls}">{w}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="warn-ok">✓ No warnings for this month</div>', unsafe_allow_html=True)


# ─── Tab 2: Residents ─────────────────────────────────────────
with tab_res:
    cfg = get_cfg()
    st.markdown("### Resident Roster")

    col1, col2, col3, col4 = st.columns([2,1,1,2])
    with col1:
        if st.button("➕ Add Resident"):
            st.session_state.res_action = "add"
            st.session_state.res_edit_id = None
    with col4:
        if st.button("🎓 Progress Year (July)", help="Advance all PGY levels by 1. PGY-4s become inactive."):
            cfg2 = get_cfg()
            graduated = []
            for r in cfg2["residents"]:
                if not r.get("active",True): continue
                if r["pgy"] >= 4: r["active"]=False; graduated.append(r["full"])
                else: r["pgy"] += 1
            save_cfg(cfg2)
            st.success(f"Year progressed! Graduated: {', '.join(graduated) if graduated else 'none'}")
            st.rerun()

    # Resident table
    rows = sorted(active_residents(cfg) + [r for r in cfg["residents"] if not r.get("active",True)],
                  key=lambda r:(r["pgy"], r["full"]))
    # Column headers for resident table
    _hc1,_hc2,_hc3,_hc4,_hc5,_hc6 = st.columns([3,1,1,3,1,1])
    _hc1.markdown("**Name**"); _hc2.markdown("**PGY**"); _hc3.markdown("**Active**")
    _hc4.markdown("**Off Service**")
    st.divider()
    for r in rows:
        active = r.get("active",True)
        # Merge old away_periods + blocked_rotations into off_service for display
        _offs = r.get("off_service", [])
        if not _offs:
            _offs = [{"start":p["start"],"end":p["end"],"reason":"Away"} for p in r.get("away_periods",[])]
            _offs += [{"start":b["start"],"end":b["end"],"reason":b.get("name","Rotation")} for b in r.get("blocked_rotations",[])]
        off_str = "; ".join(f"{o['reason']} {o['start'][:7]}→{o['end'][:7]}" for o in _offs) or "—"
        opacity = "1.0" if active else "0.45"
        with st.container():
            c1,c2,c3,c4,c5,c6 = st.columns([3,1,1,3,1,1])
            c1.markdown(f'<span style="opacity:{opacity};font-weight:500">{r["full"]}</span>',
                       unsafe_allow_html=True)
            c2.write(f"PGY-{r['pgy']}")
            c3.write("✅" if active else "⏸")
            c4.write(off_str)
            if c5.button("✎", key=f"edit_{r['id']}"):
                for _k in [f"off_svc_edit_{r['id']}"]:
                    if _k in st.session_state: del st.session_state[_k]
                st.session_state.res_action = "edit"
                st.session_state.res_edit_id = r["id"]
                st.rerun()
            if c6.button("🗑", key=f"del_{r['id']}", help="Remove resident"):
                st.session_state.res_del_id = r["id"]
                st.session_state.res_del_name = r["full"]
                st.rerun()

    # Delete confirm dialog
    if "res_del_id" in st.session_state:
        del_id   = st.session_state.res_del_id
        del_name = st.session_state.get("res_del_name", del_id)
        st.warning(f"Remove **{del_name}**?")
        dc1, dc2, dc3 = st.columns([2,2,4])
        if dc1.button("Mark Inactive", key="del_inactive"):
            cfg2 = get_cfg()
            for r in cfg2["residents"]:
                if r["id"] == del_id: r["active"] = False
            save_cfg(cfg2)
            del st.session_state["res_del_id"]
            st.session_state.pop("res_del_name", None)
            st.rerun()
        if dc2.button("Delete Permanently", key="del_perm"):
            cfg2 = get_cfg()
            cfg2["residents"] = [r for r in cfg2["residents"] if r["id"] != del_id]
            save_cfg(cfg2)
            del st.session_state["res_del_id"]
            st.session_state.pop("res_del_name", None)
            st.rerun()
        if dc3.button("Cancel", key="del_cancel"):
            del st.session_state["res_del_id"]
            st.session_state.pop("res_del_name", None)
            st.rerun()

    # Add/Edit form
    action = st.session_state.get("res_action")
    if action in ("add","edit"):
        edit_id = st.session_state.get("res_edit_id")
        existing = next((r for r in cfg["residents"] if r["id"]==edit_id), None) if edit_id else None
        st.markdown("---")
        st.markdown(f"#### {'Add' if action=='add' else 'Edit'} Resident")

        # Off Service periods — combined away + blocked rotations
        OFF_SERVICE_REASONS = ["Away","Medicine Wards","Emergency Medicine","Neurology",
                               "Emergency Psychiatry","Substance Use","Geriatric Psychiatry",
                               "Orientation","Other"]
        ofs_key = f"off_svc_{action}_{edit_id or 'new'}"
        if ofs_key not in st.session_state:
            # Load existing off_service, or migrate from old fields
            _existing_ofs = (existing or {}).get("off_service", [])
            if not _existing_ofs:
                _existing_ofs  = [{"start":p["start"],"end":p["end"],"reason":"Away"}
                                   for p in (existing or {}).get("away_periods", [])]
                _existing_ofs += [{"start":b["start"],"end":b["end"],"reason":b.get("name","Rotation")}
                                   for b in (existing or {}).get("blocked_rotations", [])]
            st.session_state[ofs_key] = list(_existing_ofs)

        st.markdown("**Off Service Periods**")
        _ofs_list = st.session_state[ofs_key]
        if _ofs_list:
            for _oi, _o in enumerate(_ofs_list):
                _ocols = st.columns([3,1])
                _ocols[0].caption(f"• {_o['reason']}: {_o['start']} → {_o['end']}")
                if _ocols[1].button("Remove", key=f"ofs_rm_{ofs_key}_{_oi}"):
                    st.session_state[ofs_key].pop(_oi)
                    st.rerun()
        else:
            st.caption("  (none)")

        st.markdown("**Add Off Service Period:**")
        _nr1, _nr2, _nr3 = st.columns([2,2,2])
        _new_reason = _nr1.selectbox("Reason", OFF_SERVICE_REASONS, key=f"ofs_rsn_{ofs_key}", label_visibility="collapsed")
        _new_dates = _nr2.date_input("Date range", value=[], key=f"ofs_dt_{ofs_key}",
                                     min_value=date(2025,1,1), max_value=date(2030,12,31),
                                     label_visibility="collapsed")
        if _nr3.button("+ Add", key=f"ofs_add_{ofs_key}"):
            if isinstance(_new_dates, (list,tuple)) and len(_new_dates) == 2:
                st.session_state[ofs_key].append({
                    "start": _new_dates[0].isoformat(),
                    "end":   _new_dates[1].isoformat(),
                    "reason": _new_reason
                })
                st.rerun()
            else:
                st.warning("Please select both a start and end date using the calendar.")

        with st.form("res_form"):
            full = st.text_input("Full Name", value="" if not existing else existing["full"])
            pgy  = st.selectbox("PGY Level", [1,2,3,4],
                                index=(existing["pgy"]-1 if existing else 0))
            active_sw = st.checkbox("Active", value=True if not existing else existing.get("active",True))
            col_s, col_c = st.columns(2)
            submitted = col_s.form_submit_button("Save", type="primary")
            cancelled = col_c.form_submit_button("Cancel")

        if cancelled:
            del st.session_state["res_action"]
            st.rerun()

        if submitted:
            if not full.strip():
                st.error("Enter a full name.")
            else:
                ofs_data = st.session_state.get(f"off_svc_{action}_{edit_id or 'new'}", [])
                cfg2 = get_cfg()
                if action=="add":
                    new_id = make_res_id(full)
                    existing_ids = {r["id"] for r in cfg2["residents"]}
                    suffix=2; base=new_id
                    while new_id in existing_ids: new_id=f"{base}_{suffix}"; suffix+=1
                    cfg2["residents"].append({"id":new_id,"full":full.strip(),"pgy":pgy,
                                              "active":True,"off_service":ofs_data})
                else:
                    for r in cfg2["residents"]:
                        if r["id"]==edit_id:
                            r["full"]=full.strip(); r["pgy"]=pgy; r["active"]=active_sw
                            r["off_service"]=ofs_data
                save_cfg(cfg2)
                del st.session_state["res_action"]
                st.rerun()


# ─── Tab 3: Holidays ─────────────────────────────────────────
with tab_hol:
    cfg = get_cfg()
    st.markdown("### Holiday Coverage")
    st.caption("Click any date to assign or edit its holiday coverage.")

    rb = res_by_id(cfg)
    all_names = [""] + [r["full"] for r in active_residents(cfg)]
    res_id_map = {r["full"]:r["id"] for r in cfg["residents"]}

    if "hol_ym" not in st.session_state: st.session_state.hol_ym = (2026,7)
    hy, hm = st.session_state.hol_ym
    c1,c2,c3 = st.columns([1,3,1])
    if c1.button("◀", key="hol_prev"):
        hm-=1
        if hm<1: hm=12; hy-=1
        st.session_state.hol_ym=(hy,hm); st.rerun()
    c2.markdown(f"<h4 style='text-align:center'>{date(hy,hm,1).strftime('%B %Y')}</h4>", unsafe_allow_html=True)
    if c3.button("▶", key="hol_next"):
        hm+=1
        if hm>12: hm=1; hy+=1
        st.session_state.hol_ym=(hy,hm); st.rerun()

    hol_lookup = {}
    for hi,hol in enumerate(cfg["holidays"]):
        for ei,entry in enumerate(hol["entries"]):
            hol_lookup[entry["date"]] = (hol["name"],entry,hi,ei)

    first_dow = date(hy,hm,1).weekday()
    num_days  = calendar.monthrange(hy,hm)[1]

    dow_hdr = st.columns(7)
    for i,dn in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]):
        dow_hdr[i].markdown(f"<div style='text-align:center;font-weight:600;font-size:0.8em'>{dn}</div>", unsafe_allow_html=True)

    for week in range(6):
        cols = st.columns(7)
        any_day = False
        for dow in range(7):
            slot = week*7+dow
            day_num = slot-first_dow+1
            col = cols[dow]
            if day_num<1 or day_num>num_days:
                col.markdown("&nbsp;", unsafe_allow_html=True)
                continue
            any_day = True
            d  = date(hy,hm,day_num)
            dk = d.isoformat()
            hinfo = hol_lookup.get(dk)
            if hinfo:
                hname_lbl,entry_lbl,_,_ = hinfo
                aptu_id  = entry_lbl.get("aptu","")
                con_id   = entry_lbl.get("consult","")
                aptu_last = rb.get(aptu_id,{}).get("full","").split()[-1] if aptu_id else ""
                con_last  = rb.get(con_id,{}).get("full","").split()[-1] if con_id else ""
                hname_short = hname_lbl[:10]
                lbl_parts = [str(day_num), hname_short]
                if aptu_last: lbl_parts.append("A: "+aptu_last)
                if con_last:  lbl_parts.append("C: "+con_last)
                lbl = "\n".join(lbl_parts)
                btype = "primary"
            else:
                lbl = str(day_num)
                btype = "secondary"
            if col.button(lbl, key=f"hd_{dk}", type=btype, use_container_width=True):
                st.session_state.hol_sel_date = dk; st.rerun()
        if not any_day: break

    st.markdown("---")
    st.markdown("#### Assign / Edit Holiday")

    sel_ds = st.session_state.get("hol_sel_date", "")
    try:
        default_date = date.fromisoformat(sel_ds) if sel_ds else date(hy,hm,1)
    except ValueError:
        default_date = date(hy,hm,1)

    ex_on_date = hol_lookup.get(default_date.isoformat())
    def_hname  = ex_on_date[0] if ex_on_date else ""
    def_aptu   = rb.get(ex_on_date[1].get("aptu",""),{}).get("full","") if ex_on_date else ""
    def_con    = rb.get(ex_on_date[1].get("consult",""),{}).get("full","") if ex_on_date else ""

    with st.form("hol_form"):
        hol_date = st.date_input("Date", value=default_date,
                                 min_value=date(2026,1,1), max_value=date(2027,12,31))
        existing_hol_names = list({h["name"] for h in cfg["holidays"]}) + ["— New holiday —"]
        def_hidx = existing_hol_names.index(def_hname) if def_hname in existing_hol_names else len(existing_hol_names)-1
        hol_name_sel = st.selectbox("Holiday group", existing_hol_names, index=def_hidx)
        new_hol_name = st.text_input("Or enter new holiday name",
                                     value="" if def_hname in existing_hol_names[:-1] else def_hname)
        def_aptu_idx = all_names.index(def_aptu) if def_aptu in all_names else 0
        def_con_idx  = all_names.index(def_con)  if def_con  in all_names else 0
        aptu_sel    = st.selectbox("APTU",                   all_names, index=def_aptu_idx)
        consult_sel = st.selectbox("Consult (blank = none)", all_names, index=def_con_idx)
        c1,c2 = st.columns(2)
        save_h = c1.form_submit_button("Assign / Update", type="primary")
        del_h  = c2.form_submit_button("Delete this date's assignment")

    if save_h:
        hname = new_hol_name.strip() or (hol_name_sel if hol_name_sel != "— New holiday —" else "")
        if not hname: st.error("Enter a holiday name.")
        else:
            cfg2=get_cfg(); dk2=hol_date.isoformat()
            aptu_id2    = res_id_map.get(aptu_sel,"")
            consult_id2 = res_id_map.get(consult_sel,"")
            entry2 = {"date":dk2,"aptu":aptu_id2,"consult":consult_id2}
            hg = next((h for h in cfg2["holidays"] if h["name"]==hname),None)
            if hg is None: cfg2["holidays"].append({"name":hname,"entries":[entry2]})
            else:
                ex2 = next((e for e in hg["entries"] if e["date"]==dk2),None)
                if ex2: ex2.update(entry2)
                else: hg["entries"].append(entry2)
            save_cfg(cfg2); st.session_state.hol_sel_date=dk2; st.rerun()

    if del_h:
        cfg2=get_cfg(); dk2=hol_date.isoformat()
        for hg in cfg2["holidays"]: hg["entries"]=[e for e in hg["entries"] if e["date"]!=dk2]
        cfg2["holidays"]=[h for h in cfg2["holidays"] if h["entries"]]
        save_cfg(cfg2)
        if "hol_sel_date" in st.session_state: del st.session_state["hol_sel_date"]
        st.rerun()

# ─── Tab 4: No-Call Requests ─────────────────────────────────
with tab_nc:
    cfg = get_cfg()
    st.markdown("### No-Call Requests")
    st.caption("Click a day to toggle the selected resident's no-call for that date. 🚫=no-call, 🔒=holiday, 📵=program no-call")

    res_options = {r["full"]:r["id"] for r in active_residents(cfg)}
    sel_res_name = st.selectbox("Resident", list(res_options.keys()), key="nc_res_sel")
    sel_res_id   = res_options[sel_res_name]

    if "nc_ym" not in st.session_state: st.session_state.nc_ym = (2026,7)
    ny, nm = st.session_state.nc_ym
    c1,c2,c3 = st.columns([1,3,1])
    if c1.button("◀", key="nc_prev"):
        nm-=1
        if nm<1: nm=12; ny-=1
        st.session_state.nc_ym=(ny,nm); st.rerun()
    c2.markdown(f"<h4 style='text-align:center'>{date(ny,nm,1).strftime('%B %Y')}</h4>", unsafe_allow_html=True)
    if c3.button("▶", key="nc_next"):
        nm+=1
        if nm>12: nm=1; ny+=1
        st.session_state.nc_ym=(ny,nm); st.rerun()

    nc_set   = {e["date"] for e in cfg.get("no_call_requests",[]) if e["resident"]==sel_res_id}
    prog_nc  = set(cfg.get("program_no_call",[]))

    first_dow = date(ny,nm,1).weekday()
    num_days  = calendar.monthrange(ny,nm)[1]

    dow_hdr = st.columns(7)
    for i,dn in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]):
        dow_hdr[i].markdown(f"<div style='text-align:center;font-weight:600;font-size:0.8em'>{dn}</div>", unsafe_allow_html=True)

    for week in range(6):
        cols = st.columns(7)
        any_day = False
        for dow in range(7):
            slot = week*7+dow
            day_num = slot-first_dow+1
            col = cols[dow]
            if day_num<1 or day_num>num_days:
                col.markdown("&nbsp;", unsafe_allow_html=True)
                continue
            any_day = True
            d   = date(ny,nm,day_num)
            dk  = d.isoformat()
            is_nc      = dk in nc_set
            is_prog_nc = dk in prog_nc
            is_hol     = bool(get_holiday(d,cfg))
            if is_nc: badge = " 🚫"
            elif is_hol: badge = " 🔒"
            elif is_prog_nc: badge = " 📵"
            else: badge = ""
            lbl = str(day_num)+badge
            btype = "primary" if is_nc else "secondary"
            if col.button(lbl, key=f"nc_{dk}_{sel_res_id}", type=btype, use_container_width=True):
                cfg2=get_cfg()
                if is_nc:
                    cfg2["no_call_requests"]=[e for e in cfg2.get("no_call_requests",[])
                                              if not (e["resident"]==sel_res_id and e["date"]==dk)]
                else:
                    cfg2.setdefault("no_call_requests",[]).append({"resident":sel_res_id,"date":dk})
                save_cfg(cfg2); st.rerun()
        if not any_day: break

    st.markdown("---")
    all_nc = sorted(e["date"] for e in cfg.get("no_call_requests",[]) if e["resident"]==sel_res_id)
    if all_nc:
        with st.expander(f"All no-call dates for {sel_res_name} ({len(all_nc)})"):
            for dk in all_nc:
                c1,c2 = st.columns([4,1])
                c1.write(f"• {date.fromisoformat(dk).strftime('%A, %B %-d, %Y')}")
                if c2.button("✕", key=f"rmnc_{dk}_{sel_res_id}"):
                    cfg2=get_cfg()
                    cfg2["no_call_requests"]=[e for e in cfg2.get("no_call_requests",[])
                                              if not (e["resident"]==sel_res_id and e["date"]==dk)]
                    save_cfg(cfg2); st.rerun()



# ─── Tab 5: Program No-Call Days ─────────────────────────────
with tab_pnc:
    cfg = get_cfg()
    st.markdown("### Program-Wide No-Call Days")
    st.caption("No one is assigned on these dates — retreats, in-service exam, graduation, etc.")

    pnc_dates = sorted(cfg.get("program_no_call",[]))

    if pnc_dates:
        for dk in pnc_dates:
            c1,c2 = st.columns([4,1])
            c1.write(f"📵  {date.fromisoformat(dk).strftime('%A, %B %-d, %Y')}")
            if c2.button("Remove", key=f"delpnc_{dk}"):
                cfg2=get_cfg()
                cfg2["program_no_call"]=[d for d in cfg2["program_no_call"] if d!=dk]
                save_cfg(cfg2); st.rerun()
    else:
        st.info("No program no-call days configured.")

    st.markdown("---")
    with st.form("pnc_form"):
        pnc_input = st.date_input("Add date", value=date(2026,9,26))
        if st.form_submit_button("➕ Add", type="primary"):
            cfg2=get_cfg(); dk2=pnc_input.isoformat()
            if dk2 not in cfg2["program_no_call"]:
                cfg2["program_no_call"].append(dk2)
                cfg2["program_no_call"].sort()
            save_cfg(cfg2); st.rerun()
